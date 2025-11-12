# source ~/barcode-env/bin/activate
# python bin/gitlab-commits-export.py --since="2025-06-01T00:00:00" --until="2025-07-01T23:59:59" --output="tangca_thang6.xlsx" --include-working-hours --branch=dev/nam-webform --keyword fix bug urgent

import argparse
import subprocess
import json
from datetime import datetime, time
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side

import re

print("🔥 Đang chạy file:", __file__)

# Hàm cắt chuỗi lấy phần trước dấu '|'
def cut_after_pipe(s):
    return s.split('|', 1)[0].strip()

# Hàm tìm OP id trong message dạng {OP#1234}
def extract_op_id(message):
    m = re.search(r"\{OP#(\d+)\}", message)
    return m.group(1) if m else None

def build_branch_index():
    """
    Xây dựng ánh xạ commit → danh sách branch chứa nó.
    Dùng `git for-each-ref` để lấy thông tin nhanh chóng.
    """
    try:
        result = subprocess.run(
            ["git", "for-each-ref", "--format=%(objectname)|%(refname:short)", "refs/heads", "refs/remotes/origin"],
            stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True, check=True, encoding='utf-8'
        )
        lines = result.stdout.strip().splitlines()
        hash_to_branch = defaultdict(list)
        for line in lines:
            parts = line.split("|", 1)
            if len(parts) == 2:
                commit_hash, branch_name = parts
                hash_to_branch[commit_hash].append(branch_name)
        return hash_to_branch
    except Exception as e:
        print("❌ Lỗi khi build branch index:", e)
        return {}
    
# Cache kết quả tìm branch cho commit hash để tránh gọi git nhiều lần
_branch_cache = {}
_branch_index = build_branch_index()

def get_branch_for_hash(commit_hash):
    if commit_hash in _branch_cache:
        return _branch_cache[commit_hash]

    # 1. Kiểm tra trong index
    if commit_hash in _branch_index:
        branches = _branch_index[commit_hash]
        branch_str = ", ".join(branches)
        _branch_cache[commit_hash] = branch_str
        return branch_str

    # 2. Nếu không có trong index, fallback dùng git
    try:
        result = subprocess.run(
            ["git", "branch", "--contains", commit_hash, "--format=%(refname:short)"],
            stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True, check=True, encoding='utf-8'
        )
        branches = [line.strip() for line in result.stdout.splitlines() if line.strip()]
        branch_str = ", ".join(branches)

        # Lưu cache dù kết quả rỗng
        _branch_cache[commit_hash] = branch_str
        return branch_str
    except subprocess.CalledProcessError as e:
        print(f"⚠️ Lỗi khi chạy git branch --contains {commit_hash}: {e}")
    except Exception as e:
        print(f"⚠️ Lỗi không xác định khi tìm branch cho {commit_hash}: {e}")

    # fallback cuối cùng nếu lỗi
    _branch_cache[commit_hash] = ""
    return ""

# Lấy danh sách commit con của một merge commit theo author (lọc bằng --author)
def get_child_commits_by_author(parent, merge_hash, author):
    """
    Lấy danh sách commit con trong phạm vi parent..merge_hash của author.
    parent: hash parent (thường là parent thứ 2 của merge commit)
    merge_hash: hash merge commit
    author: tên tác giả lọc theo --author
    """
    try:
        result = subprocess.run(
            ["git", "log", f"{parent}..{merge_hash}", "--pretty=format:%h|%an", f"--author={author}"],
            stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True, check=True, encoding='utf-8'
        )
        lines = result.stdout.strip().splitlines()
        hashes = [line.split("|")[0] for line in lines if line.strip()]
        return hashes
    except Exception:
        return []

# Hàm in báo cáo ra console
def print_console_report(data, count_stat, include_working_hours):
    BOLD = '\033[1m'
    RESET = '\033[0m'
    for author, dates in data.items():
        print("="*60)
        print(f"{BOLD}👤 Tác giả: {author}{RESET}")
        print("="*60)
        for date, commits in sorted(dates.items()):
            print(f"📅 Ngày: {date.strftime('%d/%m/%Y')}")
            for c in commits:
                print(f"  🕒 {c['time']} - 🔗 {c['hash']} | {c['type']} | [{c['branch']}] | {c['message']}")
                if c['child_commits']:
                    print(f"     ↳ Gồm commit con: {', '.join(c['child_commits'])}")
        o = count_stat[author]["overtime"]
        w = count_stat[author]["working"]
        summary = f"🔥 Tăng ca: {o} commit(s)"
        if include_working_hours:
            summary += f" | ⏰ Trong giờ: {w} commit(s)"
        print(f"\n📊 {summary}")
        print("."*60)

# --- Phần xử lý chính ---

# Định nghĩa tham số đầu vào cho script
parser = argparse.ArgumentParser(description="Báo cáo commit tăng ca: hiển thị + xuất Excel")
parser.add_argument("--author", help="Tác giả (tuỳ chọn)")
parser.add_argument("--since", default="2025-06-01T00:00:00", help="2025-06-01T00:00:00")
parser.add_argument("--until", default="2025-06-30T23:59:59", help="2025-06-01T00:00:00")
parser.add_argument("--output", default="commits.xlsx", help="commits.xlsx")
parser.add_argument("--include-working-hours", action="store_true", help="Bao gồm cả commit trong giờ hành chính")
parser.add_argument("--branch", help="Chỉ lấy commit thuộc nhánh cụ thể (ví dụ: develop, feature/x)")
parser.add_argument("--keyword", nargs="+", help="Lọc commit chứa từ khóa trong message (vd: fix, bug, urgent)")

args = parser.parse_args()

# Tạo command git log với format JSON để dễ parse
git_cmd = ["git", "log"]

# Nếu có --branch thì chỉ lấy log của nhánh đó, ngược lại lấy tất cả
if args.branch:
    git_cmd.append(args.branch)
else:
    git_cmd.append("--all")

git_cmd += [
    f"--since={args.since}",
    f"--until={args.until}",
    '--pretty=format:%ad\x1f%H\x1f%an\x1f%D\x1f%s\x1f%P',
    "--date=format:%Y-%m-%d %H:%M:%S"
]

if args.author:
    git_cmd.append(f"--author={args.author}")

# Chạy lệnh git log và lấy output theo định dạng phân cách đặc biệt
# result = subprocess.run(git_cmd, capture_output=True, text=True)
result = subprocess.run(git_cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True, encoding='utf-8')

# Parse từng dòng thủ công, dùng dấu phân cách \x1f (Unit Separator)
entries = []
for line in result.stdout.strip().splitlines():
    parts = line.strip().split("\x1f")
    if len(parts) == 6:
        date, hash_, author, ref, msg, parents = parts
        entries.append({
            "date": date.strip(),
            "hash": hash_.strip(),
            "author": author.strip(),
            "ref": ref.strip(),
            "msg": msg.strip(),
            "parents": parents.strip()
        })

# Dữ liệu lưu commit theo cấu trúc: author -> ngày -> list commit
data = defaultdict(lambda: defaultdict(list))
# Thống kê số commit tăng ca và trong giờ
count_stat = defaultdict(lambda: {"overtime": 0, "working": 0})

# Định nghĩa giờ làm việc chuẩn
start_work = time(8, 0)
end_work = time(17, 30)

# Duyệt từng commit trong danh sách
for entry in entries:
    try:
        dt = datetime.strptime(entry["date"], "%Y-%m-%d %H:%M:%S")

        # Bỏ qua nếu không khớp từ khóa (nếu có chỉ định)
        if args.keyword:
            msg_lower = entry["msg"].lower()
            if not any(kw.lower() in msg_lower for kw in args.keyword):
                continue

        # Xác định commit ngoài giờ nếu giờ commit không nằm trong khung làm việc
        is_overtime = not (start_work <= dt.time() <= end_work)

        # Nếu không muốn include commit trong giờ làm thì bỏ qua
        if not args.include_working_hours and not is_overtime:
            continue

        # Lấy tên author, lọc bỏ phần sau dấu '|'
        author = cut_after_pipe(entry["author"])

        # Lấy thông tin ref (nhánh, tag,...)
        ref_info = entry["ref"].strip()

        # Mã commit hash
        commit_hash = entry["hash"].strip()

        # Lấy danh sách cha của commit (để xác định merge)
        parents = entry.get("parents", "").strip().split()

        # Xác định commit merge hay bình thường
        is_merge = len(parents) > 1

        # Lấy commit con của merge do author này tạo ra (nếu là merge)
        child_commits = []
        if is_merge and len(parents) > 1:
            # Lấy parent thứ 2 làm điểm bắt đầu phạm vi log commit con
            child_commits = get_child_commits_by_author(parents[1], commit_hash, author)

        # Xác định type commit (merge-with-contribution, merge-only, commit)
        commit_type = (
            "merge-with-contribution" if is_merge and child_commits else
            "merge-only" if is_merge else
            "commit"
        )

        # Xác định branch theo thứ tự ưu tiên:
        branch = ""
        m = re.search(r'HEAD -> ([^,\s]+)', ref_info)
        if m:
            branch = m.group(1)
        elif "origin/" in ref_info:
            m = re.search(r'origin/([^,\s]+)', ref_info)
            if m:
                branch = m.group(1)
        elif ref_info:
            branch = ref_info.split(",")[0].strip()

        # Nếu vẫn chưa có branch, gọi git để lấy
        if not branch:
            branch = get_branch_for_hash(commit_hash)

        # Chuẩn hóa lại tên hiển thị branch
        branch = re.sub(r"^origin/", "", branch)

        # Lưu dữ liệu commit vào dict theo author và ngày
        data[author][dt.date()].append({
            "time": dt.strftime("%H:%M"),
            "hash": commit_hash,
            "message": entry["msg"].strip(),
            "branch": branch,
            "event": "merge" if is_merge else "commit",
            "type": commit_type,
            "child_commits": child_commits,
            "is_overtime": is_overtime
        })

        # Cập nhật thống kê tăng ca hoặc trong giờ
        if is_overtime:
            count_stat[author]["overtime"] += 1
        else:
            count_stat[author]["working"] += 1

    except Exception as e:
        print(f"❗️Lỗi xử lý commit {entry.get('hash', 'unknown')} ({entry.get('msg', '')[:50]}...): {e}")
        continue

# In báo cáo ra console
print_console_report(data, count_stat, args.include_working_hours)

# ------- Xuất Excel -------
thick_top = Border(top=Side(style='thin'))
thin_top = Border(top=Side(style='thin'))  # hoặc style='medium' nếu bạn muốn đậm hơn

wb = Workbook()
ws = wb.active
ws.title = "Tăng ca"

headers = ["Tác giả", "Ngày", "Thời gian", "Mã commit", "Event", "Type", "Branch", "Nội dung", "OP Link", "Child Commits"]
ws.append(headers)

header_fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")  # Cam nhạt
summary_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
highlight_fill = PatternFill(start_color="FFC07F", end_color="FFC07F", fill_type="solid")  # Lavender
alt_row_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")  # Nền xen kẽ

for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = Font(bold=True)
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

ws.freeze_panes = "A2"
row = 2
first_author = True
for author, dates in data.items():
    if not first_author:
        # Chỉ thêm dòng trống nếu không phải tác giả đầu tiên
        ws.append([""] * len(headers))
        row += 1
    else:
        first_author = False

    author_start_row = row

    for date, commits in sorted(dates.items()):
        date_start_row = row
        for c in commits:
            op_id = extract_op_id(c["message"])
            row_data = [
                author,
                date.strftime("%d/%m/%Y"),
                c["time"],
                c["hash"],
                c["event"],
                c["type"],
                c["branch"],
                c["message"],
                op_id if op_id else "",
                ", ".join(c["child_commits"]) if c["child_commits"] else ""
            ]
            ws.append(row_data)
            # Tô màu xen kẽ cho dòng chẵn, nhưng chỉ nếu không phải dòng merged hoặc tổng kết
            if row_data[0] != "" and row_data[1] != "":  # Không phải dòng merge
                if row % 2 == 0:
                    for col_idx in range(1, len(headers) + 1):
                        ws.cell(row=row, column=col_idx).fill = alt_row_fill


            # Tạo hyperlink nếu có OP link
            if op_id:
                ws.cell(row=row, column=9).hyperlink = f"https://work.fsofts.com/work_packages/{op_id}"
                ws.cell(row=row, column=9).style = "Hyperlink"

            # Canh chỉnh text trong các ô
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row, column=col_idx).alignment = Alignment(vertical="top", horizontal="left")

            # In đậm cột Event và Type nếu là merge
            if c["event"] == "merge":
                for col_idx in (5, 6):  # cột E và F
                    ws.cell(row=row, column=col_idx).font = Font(bold=True)

            # Tô màu cột thời gian nếu là commit tăng ca
            if c["is_overtime"]:
                ws.cell(row=row, column=3).fill = highlight_fill

            row += 1

        # # Gộp ô cho cột ngày nếu nhiều commit trong ngày
        # date_end_row = row - 1
        # if date_end_row > date_start_row:
        #     ws.merge_cells(start_row=date_start_row, end_row=date_end_row, start_column=2, end_column=2)
        #     for r in range(date_start_row, date_end_row + 1):
        #         ws.cell(r, 2).alignment = Alignment(horizontal="center", vertical="top")
        # else:
        #     ws.cell(date_start_row, 2).alignment = Alignment(horizontal="center", vertical="top")
        
        # Chỉ điền giá trị ngày ở dòng đầu tiên, các dòng còn lại để trống
        for r in range(date_start_row + 1, row):
            ws.cell(r, 2).value = ""

    # # Gộp ô cho cột tác giả nếu nhiều ngày
    # author_end_row = row - 1
    # if author_end_row > author_start_row:
    #     ws.merge_cells(start_row=author_start_row, end_row=author_end_row, start_column=1, end_column=1)
    #     for r in range(author_start_row, author_end_row + 1):
    #         for col in range(1, 11):  # Cột A → J (1 → 10)
    #             cell = ws.cell(r, col)
    #             if col == 1:
    #                 cell.alignment = Alignment(horizontal="left", vertical="top")
    #                 cell.font = Font(bold=True)
    #             if r == author_start_row:
    #                 cell.border = thick_top
    
    # Ghi tên tác giả ở dòng đầu tiên, các dòng còn lại để trống
    for r in range(author_start_row, row):
        cell = ws.cell(r, 1)
        if r == author_start_row:
            cell.value = author
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="left", vertical="top")
            cell.fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")

            # Border top cho cả hàng
            for col_idx in range(1, len(headers) + 1):
                ws.cell(r, col_idx).border = Border(top=Side(style='medium'))
        else:
            cell.value = ""


    # Thêm dòng tổng kết số commit
    o = count_stat[author]["overtime"]
    w = count_stat[author]["working"]
    summary = f"Tăng ca: {o} commits"
    if args.include_working_hours:
        summary += f" | Trong giờ: {w} commits"

    full_summary = f"Tổng cộng: {summary}"
    ws.append([full_summary])

    # Gộp dòng tổng kết từ cột A → J
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=10)
    merged_cell = ws.cell(row=row, column=1)
    merged_cell.alignment = Alignment(horizontal="left", vertical="top")
    merged_cell.fill = summary_fill
    merged_cell.font = Font(bold=True)

    row += 1

# Tự động chỉnh lại chiều rộng cột, giới hạn tối đa 30 ký tự
MAX_WIDTH = 30
for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col[1:]:  # Bỏ header
        try:
            cell_value = str(cell.value or "")
        except:
            cell_value = ""
        if len(cell_value) > max_len:
            max_len = len(cell_value)
    max_len = min(max_len, MAX_WIDTH)
    ws.column_dimensions[col_letter].width = max(max_len + 1, 8)

# Lưu file Excel
wb.save(args.output)
print(f"\n✅ Đã xuất file Excel: {args.output}")
